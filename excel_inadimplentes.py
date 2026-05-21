"""
excel_inadimplentes.py
======================

Leitor de planilha Excel exportada do painel da Efí (Receber > Gestão de
cobranças > Boletos > Exportar).

Suporta tanto .xlsx (openpyxl) quanto .xls legado (xlrd==1.2.0).

Estrutura esperada da planilha (cabeçalho na linha 1):

    Loja | Número da Cobrança | Valor (R$) | Status | Data de Emissão |
    Forma de Cobrança | Itens | Nome | E-mail | Telefone | Documento |
    Data de Vencimento | Valor Pago | Pago Em

Apenas linhas com Status = "Inadimplente" são consideradas. Status
"Cancelado", "Pago", "Aguardando" e "Marcado como pago" são ignorados.

Saída: lista de dicts no MESMO formato usado pelo efi_client.py para que
o relatorio_inadimplentes.py possa consumir indistintamente:

    {
        "cnpj": "12345678000100",         # só dígitos
        "nome": "Razão social",
        "valor": 990.00,                  # float em reais
        "vencimento": "2026-05-20",       # ISO YYYY-MM-DD
        "dias_atraso": 1,
        "numero_cobranca": "1013504771",
        "fonte": "excel",
    }
"""

from __future__ import annotations

import re
from datetime import datetime, date
from pathlib import Path
from typing import Iterable, List, Dict, Any, Optional, Tuple

STATUS_INADIMPLENTES = {"inadimplente", "vencido", "atrasado"}


def _so_digitos(valor: Any) -> str:
    if valor is None:
        return ""
    return re.sub(r"\D", "", str(valor))


def _para_float(valor: Any) -> float:
    if valor is None or valor == "":
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip()
    s = s.replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _para_iso(valor: Any) -> Optional[str]:
    if valor is None or valor == "":
        return None
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%Y-%m-%d")
    s = str(valor).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _dias_atraso(venc_iso: Optional[str]) -> int:
    if not venc_iso:
        return 0
    try:
        v = datetime.strptime(venc_iso, "%Y-%m-%d").date()
    except ValueError:
        return 0
    delta = (date.today() - v).days
    return max(0, delta)


def _ler_planilha(caminho: Path) -> Tuple[List[str], List[List[Any]]]:
    """Lê a planilha e devolve (header, linhas). Suporta .xls e .xlsx."""
    ext = caminho.suffix.lower()
    if ext == ".xlsx":
        try:
            import openpyxl
        except ImportError as e:
            raise RuntimeError("openpyxl não instalado. Rode: pip install openpyxl") from e
        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return [], []
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        data = [list(r) for r in rows[1:]]
        return header, data
    elif ext == ".xls":
        try:
            import xlrd
        except ImportError as e:
            raise RuntimeError("xlrd==1.2.0 não instalado. Rode: pip install xlrd==1.2.0") from e
        wb = xlrd.open_workbook(str(caminho))
        ws = wb.sheet_by_index(0)
        if ws.nrows == 0:
            return [], []
        header = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
        data: List[List[Any]] = []
        for r in range(1, ws.nrows):
            linha: List[Any] = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                # xlrd type 3 = date
                if cell.ctype == 3:
                    try:
                        y, m, d, *_ = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                        linha.append(date(y, m, d))
                    except Exception:
                        linha.append(cell.value)
                else:
                    linha.append(cell.value)
            data.append(linha)
        return header, data
    else:
        raise ValueError(f"Extensão não suportada: {ext} (use .xls ou .xlsx)")


def carregar_inadimplentes(caminho: str | Path) -> List[Dict[str, Any]]:
    """Lê a planilha e devolve apenas inadimplentes normalizados."""
    caminho = Path(caminho)
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {caminho}")

    header, linhas = _ler_planilha(caminho)
    if not header:
        return []

    def idx(nome: str) -> Optional[int]:
        nome_low = nome.lower()
        for i, h in enumerate(header):
            if h.lower().strip() == nome_low:
                return i
        return None

    i_loja = idx("Loja")
    i_num = idx("Número da Cobrança")
    i_valor = idx("Valor (R$)")
    i_status = idx("Status")
    i_nome = idx("Nome")
    i_email = idx("E-mail")
    i_tel = idx("Telefone")
    i_doc = idx("Documento")
    i_venc = idx("Data de Vencimento")

    out: List[Dict[str, Any]] = []
    for row in linhas:
        if row is None or all(c is None or c == "" for c in row):
            continue
        status = ""
        if i_status is not None and row[i_status]:
            status = str(row[i_status]).strip().lower()
        if status not in STATUS_INADIMPLENTES:
            continue

        cnpj = _so_digitos(row[i_doc] if i_doc is not None else "")
        if not cnpj:
            continue

        venc_iso = _para_iso(row[i_venc]) if i_venc is not None else None
        nome = str(row[i_nome]).strip() if i_nome is not None and row[i_nome] else ""
        valor = _para_float(row[i_valor]) if i_valor is not None else 0.0
        num = str(row[i_num]).strip() if i_num is not None and row[i_num] else ""

        out.append({
            "cnpj": cnpj,
            "nome": nome,
            "valor": valor,
            "vencimento": venc_iso,
            "dias_atraso": _dias_atraso(venc_iso),
            "numero_cobranca": num,
            "loja": str(row[i_loja]).strip() if i_loja is not None and row[i_loja] else "",
            "email": str(row[i_email]).strip() if i_email is not None and row[i_email] else "",
            "telefone": str(row[i_tel]).strip() if i_tel is not None and row[i_tel] else "",
            "fonte": "excel",
        })

    return out


def consolidar_por_cnpj(itens: Iterable[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """Agrupa boletos pelo CNPJ — soma valor_aberto e pega dias_max."""
    agreg: Dict[str, Dict[str, Any]] = {}
    for c in itens:
        cnpj = c["cnpj"]
        entry = agreg.setdefault(cnpj, {
            "cnpj": cnpj,
            "nome": c.get("nome", ""),
            "boletos": [],
            "total_aberto": 0.0,
            "dias_max": 0,
        })
        entry["boletos"].append(c)
        entry["total_aberto"] += float(c.get("valor") or 0.0)
        entry["dias_max"] = max(entry["dias_max"], int(c.get("dias_atraso") or 0))
        if not entry["nome"]:
            entry["nome"] = c.get("nome", "")
    return agreg


if __name__ == "__main__":
    import sys, json
    if len(sys.argv) < 2:
        print("uso: python excel_inadimplentes.py arquivo.xls(x)")
        sys.exit(1)
    inad = carregar_inadimplentes(sys.argv[1])
    print(f"[excel_inadimplentes] {len(inad)} inadimplentes carregados")
    print(json.dumps(consolidar_por_cnpj(inad), indent=2, ensure_ascii=False, default=str)[:3000])
